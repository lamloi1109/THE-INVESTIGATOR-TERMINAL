"""Build portfolio-cms-seed.xlsx from the 5 seed CSVs in this folder.

Each CSV → one sheet (tab) with the exact name the schema expects:
Profile, Projects, Experience, TechStack, Education.

Header row is bold. Column widths auto-fit (capped at 80). Multi-line
cells (problem/solution/result/achievements/highlights) wrap text so
the user can read them in Sheets after paste.
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

HERE = Path(__file__).parent
OUT = HERE / "portfolio-cms-seed.xlsx"

SHEETS = [
    ("Profile",    "profile.csv"),
    ("Projects",   "projects.csv"),
    ("Experience", "experience.csv"),
    ("TechStack",  "tech-stack.csv"),
    ("Education",  "education.csv"),
]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F2937")
BODY_FONT = Font(name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()
wb.remove(wb.active)

for tab_name, csv_name in SHEETS:
    ws = wb.create_sheet(tab_name)
    rows = list(csv.reader((HERE / csv_name).open(encoding="utf-8")))
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = BODY_FONT
                cell.alignment = WRAP
    ws.row_dimensions[1].height = 22
    # Column widths: header length × 1.2, capped to [12, 60]
    for c, header in enumerate(rows[0], start=1):
        from openpyxl.utils import get_column_letter
        width = max(12, min(60, int(len(header) * 1.4)))
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {OUT} with {len(SHEETS)} sheets")
